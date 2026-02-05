from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
import pandas as pd
import random

def base_role(role_name):
        return role_name.split(" ")[0] 

def resident_score_expr(assign, days, day_roles, resident, weekend_days, weekend_rounds_df, ns_residents_df):
    """
    Build the linear expression for a resident's score:
    - 1 point per weekday shift
    - 2 points per weekend shift (based on weekend_days)
    - +2 points per WR date (WR counted per day)
    - +2 NS bonus if resident in ns_residents_df

    `days` is an iterable of dates used as keys in `assign` (will be normalized).
    `assign[(d, role, resident)]` must be valid for d in days, role in day_roles.
    """
    # Normalize resident name for comparisons
    resident_key = str(resident).strip().lower()

    # Normalize days to Timestamp objects (and build a set for fast membership checks)
    normalized_days = []
    for d in days:
        nd = pd.Timestamp(d).normalize()
        normalized_days.append(nd)
    days_set = set(normalized_days)

    # Base score: weighted by weekday/weekend
    # Note: use the original `days` iteration if assign keys expect the same objects;
    # we use normalized_days to match weekend_days which should also be normalized.
    expr = sum(
        assign[(d, role, resident)] * (2 if pd.Timestamp(d).normalize() in weekend_days else 1)
        for d in days for role in day_roles
    )

    # Weekend round (WR) bonus: +2 per WR day that falls inside our days window
    if weekend_rounds_df is not None and not weekend_rounds_df.empty:
        # Build cleaned names (strip + lower) and normalized dates
        wr_df = weekend_rounds_df.copy()
        # Ensure 'name' and 'date' exist
        if "name" in wr_df.columns and "date" in wr_df.columns:
            # Normalize name column for comparison
            wr_df["__name_clean"] = wr_df["name"].astype(str).str.strip().str.lower()
            # Normalize date column to pd.Timestamp normalized
            wr_df["__date_norm"] = pd.to_datetime(wr_df["date"]).dt.normalize()

            # Filter rows for this resident
            resident_wr_dates = set(
                wr_df.loc[wr_df["__name_clean"] == resident_key, "__date_norm"].tolist()
            )

            # Count only WR days that are inside the schedule days we consider
            wr_days_in_window = resident_wr_dates.intersection(days_set)

            # Add 2 points per WR day
            expr += 2 * len(wr_days_in_window)

    # NS bonus: +2 if in ns_residents_df
    if ns_residents_df is not None and not ns_residents_df.empty:
        if "name" in ns_residents_df.columns:
            ns_names = set(ns_residents_df["name"].astype(str).str.strip().str.lower())
            if resident_key in ns_names:
                expr += 2
        else:
            # fallback if ns_residents_df is just a list-like of names
            try:
                ns_names = set([str(x).strip().lower() for x in ns_residents_df])
                if resident_key in ns_names:
                    expr += 2
            except Exception:
                pass

    return expr


# ==========================================================
# ================== PREPROCESS DATA =======================
# ==========================================================

def rotate_list(lst, offset):
    return lst[offset % len(lst):] + lst[:offset % len(lst)] if lst else []

# Helper: expand date ranges from NF column
def expand_dates(date_range_str, base_year, anchor_month=None):
    """
    Expand date ranges like "20-29 Dec", "28 Dec - 3 Jan", "3-9 Jan"
    into explicit dates, inferring rollover years automatically.

    Parameters
    ----------
    date_range_str : str
        The range string (e.g. "28 Dec - 3 Jan").
    base_year : int
        The year of the display_start_date.
    anchor_month : int, optional
        Month number of the calendar start (default = 12 for December).
        If display_start_date is January, pass anchor_month=1.
    """
    import pandas as pd
    from datetime import timedelta
    import re

    if not date_range_str or str(date_range_str).strip().lower() == "no":
        return []

    if anchor_month is None:
        anchor_month = 12  # default to December if not provided

    s = str(date_range_str).replace("–", "-").replace("—", "-").strip()
    s = re.sub(r"\s*-\s*", "-", s)
    parts = [p.strip() for p in re.split(r"\band\b", s, flags=re.IGNORECASE) if p.strip()]

    def month_num(month_name, year_hint):
        return pd.Timestamp(f"1 {month_name} {year_hint}").month

    def infer_year_for_month(mnum):
        # If calendar starts in January
        if anchor_month == 1:
            if mnum in (1, 2):   # Jan, Feb
                return base_year
            elif mnum == 12:     # Dec
                return base_year - 1
            else:
                return base_year
        # If calendar starts in December
        elif anchor_month == 12:
            if mnum == 12:
                return base_year
            elif mnum in (1, 2):
                return base_year + 1
            else:
                return base_year
        else:
            return base_year

    dates_out = []

    for part in parts:
        if "-" not in part:
            # Single date
            pieces = part.split()
            if len(pieces) < 2:
                continue
            day = int(pieces[0])
            month = pieces[1]
            mnum = month_num(month, base_year)
            year = infer_year_for_month(mnum)
            dt = pd.Timestamp(year=year, month=mnum, day=day).normalize()
            dates_out.append(dt)
            continue

        # Range
        start_str, end_str = [x.strip() for x in part.split("-", 1)]
        start_parts = start_str.split()
        end_parts   = end_str.split()

        # Start
        start_day = int(start_parts[0])
        start_month = start_parts[1] if len(start_parts) >= 2 else (
            end_parts[1] if len(end_parts) >= 2 else None
        )
        if not start_month:
            continue
        start_mnum = month_num(start_month, base_year)
        start_year = infer_year_for_month(start_mnum)
        start_dt = pd.Timestamp(year=start_year, month=start_mnum, day=start_day)

        # End
        end_day = int(end_parts[0])
        end_month = end_parts[1] if len(end_parts) >= 2 else start_month
        end_mnum = month_num(end_month, base_year)
        end_year = infer_year_for_month(end_mnum)

        # Internal rollover (Dec → Jan)
        if end_year == start_year and end_mnum < start_mnum:
            end_year = start_year + 1

        end_dt = pd.Timestamp(year=end_year, month=end_mnum, day=end_day)

        # Expand
        for i in range((end_dt - start_dt).days + 1):
            dates_out.append((start_dt + timedelta(days=i)).normalize())

    return dates_out

def calculate_max_limits(residents, nf_residents, resident_max_limit, nf_max_limit, night_counts):
    """
    Calculate maximum shifts, points, and weekend limits for each resident.
    A resident is treated as NF only if they have more than 2 NF nights.
    """
    max_shifts, max_points, weekend_limits = {}, {}, {}

    for resident in residents:

        # Decide if resident should be classified as NF
        is_real_nf = (resident in nf_residents) and (night_counts.get(resident, 0) > 2)

        if is_real_nf:
            # Use NF limits
            shifts, points, weekend = nf_max_limit
        else:
            # Use normal limits
            shifts, points, weekend = resident_max_limit

        # Store
        max_shifts[resident] = shifts
        max_points[resident] = points
        weekend_limits[resident] = weekend

    return max_shifts, max_points, weekend_limits


def build_fixed_preassigned(nf_calendar_df,
                            ns_residents,
                            weekend_rounds_df,
                            preassigned_ns_df=None,
                            preassigned_wr_df=None):

    fixed = {}

    def ts(x):
        return pd.to_datetime(x)

    # ---- 1. NF Calendar ----
    for _, row in nf_calendar_df.iterrows():
        for role_col in nf_calendar_df.columns:
            if role_col.lower() in ["date", "day", "weekday"]:
                continue
            r = row[role_col]
            if pd.isna(r):
                continue
            fixed.setdefault(r.strip(), set()).add(ts(row["date"]))

    # ---- 2. NS ----
    for _, row in ns_residents.iterrows():
        r = row["name"].strip()
        fixed.setdefault(r, set()).add(ts(row["date"]))

    # ---- 3. WR ----
    for _, row in weekend_rounds_df.iterrows():
        r = row["name"].strip()
        fixed.setdefault(r, set()).add(ts(row["date"]))

    # ---- 4. Excel NS ----
    if preassigned_ns_df is not None:
        for _, row in preassigned_ns_df.iterrows():
            r = row["name"].strip()
            fixed.setdefault(r, set()).add(ts(row["date"]))

    # ---- 5. Excel WR ----
    if preassigned_wr_df is not None:
        for _, row in preassigned_wr_df.iterrows():
            r = row["name"].strip()
            fixed.setdefault(r, set()).add(ts(row["date"]))

    return fixed

def get_all_assigned_dates(
    r,
    solver,
    assign,
    days,
    roles,
    ns_df=None,
    wr_df=None,
    nf_calendar_df=None,
    extra_preassigned=None
):
    """
    Collect ALL assigned dates for a resident from:
        - OR-Tools solver assignments
        - NS preassignments (ns_df)
        - WR preassignments (wr_df)
        - NF calendar (nf_calendar_df)
        - extra_preassigned: list/dict of additional custom assignments

    Returns:
        A sorted list of unique datetime.date objects.
    """

    all_dates = set()

    # ----------------------------------------------------
    # 1) OR-TOOLS ASSIGNED SHIFTS
    # ----------------------------------------------------
    for d in days:
        date_val = d.date() if hasattr(d, "date") else d
        for role in roles:
            try:
                if solver.Value(assign[(d, role, r)]) == 1:
                    all_dates.add(date_val)
            except Exception:
                pass

    # ----------------------------------------------------
    # 2) NS PREASSIGNMENTS
    # ----------------------------------------------------
    if ns_df is not None and isinstance(ns_df, pd.DataFrame) and not ns_df.empty:
        if "name" in ns_df.columns and "date" in ns_df.columns:
            ns_dates = ns_df.loc[ns_df["name"] == r, "date"]
            for d in ns_dates:
                all_dates.add(pd.to_datetime(d).date())

    # ----------------------------------------------------
    # 3) WR PREASSIGNMENTS
    # ----------------------------------------------------
    if wr_df is not None and isinstance(wr_df, pd.DataFrame) and not wr_df.empty:
        if "name" in wr_df.columns and "date" in wr_df.columns:
            wr_dates = wr_df.loc[wr_df["name"] == r, "date"]
            for d in wr_dates:
                all_dates.add(pd.to_datetime(d).date())

    # ----------------------------------------------------
    # 4) NF CALENDAR (night float)
    # ----------------------------------------------------
    if nf_calendar_df is not None and isinstance(nf_calendar_df, pd.DataFrame) and not nf_calendar_df.empty:
        if "name" in nf_calendar_df.columns and "date" in nf_calendar_df.columns:
            nf_dates = nf_calendar_df.loc[nf_calendar_df["name"].strip() == r, "date"]
            for d in nf_dates:
                all_dates.add(pd.to_datetime(d).date())

    # ----------------------------------------------------
    # 5) EXTRA CUSTOM PREASSIGNMENTS
    # ----------------------------------------------------
    if extra_preassigned:
        # expects either dict: {resident: [dates]}
        # OR list of (resident, date)
        if isinstance(extra_preassigned, dict):
            for d in extra_preassigned.get(r, []):
                all_dates.add(pd.to_datetime(d).date())

        elif isinstance(extra_preassigned, list):
            for item in extra_preassigned:
                if isinstance(item, tuple) and len(item) == 2:
                    name, d = item
                    if name == r:
                        all_dates.add(pd.to_datetime(d).date())

    # ----------------------------------------------------
    # Return sorted list
    # ----------------------------------------------------
    return sorted(all_dates)

def read_excel_as_displayed(path, sheet_name=0):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for h, cell in zip(headers, row):
            if cell.is_date and cell.value is not None:
                # Format all Excel dates as '23 Nov'
                value = cell.value.strftime("%d %b")
            else:
                value = cell.value
            row_data[h] = value
        data.append(row_data)

    return pd.DataFrame(data)

# DEBUG
def assert_penalties(name, penalties):
    if penalties is None:
        raise ValueError(f"{name} returned None")
    if not isinstance(penalties, list):
        raise TypeError(f"{name} returned {type(penalties)}")
    for p in penalties:
        if not hasattr(p, "Proto"):
            raise TypeError(f"{name} contains non-var: {p}")

    print(f"✅ {name}: {len(penalties)} penalties")
