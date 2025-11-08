import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# -------------------------
# Save Schedule DataFrame
# -------------------------
def save_schedule_as_excel(df, output_path="Schedule.xlsx"):

    df = df.fillna("")
    # Convert 'Date' column to readable format
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%d-%b')


    # -------------------------
    # Reorder columns: make ER sections Day/Night pairs and move WR last
    # -------------------------
    day_night_cols = []
    section_prefixes = ['er-1', 'er-2', 'er']  # adjust as needed
    for sec in section_prefixes:
        day_col = f"{sec} day"
        night_col = f"{sec} night"
        if day_col in df.columns:
            day_night_cols.append(day_col)
        if night_col in df.columns:
            day_night_cols.append(night_col)

    # Keep other columns
    other_cols = [c for c in df.columns if c not in day_night_cols + ['wr']]
    # Reorder: other columns + day/night pairs + WR at the end
    df = df[other_cols + day_night_cols + ['wr']]

    # Convert 'Date' column to readable format
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%d-%b')

    # -------------------------
    # Split WR column into two if multiple names
    # -------------------------
    if "wr" in df.columns:
        # Detect if any row contains a comma (meaning multiple names)
        has_multiple = df["wr"].astype(str).str.contains(",").any()

        if has_multiple:
            wr_split = df["wr"].fillna("").str.split(",", n=1, expand=True)
            wr_split.columns = ["wr-1", "wr-2"]

            # Clean up spaces
            wr_split["wr-1"] = wr_split["wr-1"].str.strip()
            wr_split["wr-2"] = wr_split["wr-2"].str.strip()

            # Replace original WR column with the split ones
            df = pd.concat([df.drop(columns=["wr"]), wr_split], axis=1)
        else:
            # Rename to wr-1 for consistent naming if only single WR per day
            df = df.rename(columns={"wr": "wr-1"})


    df.to_excel(output_path, index=False, sheet_name='Schedule')
    wb = load_workbook(output_path)
    ws = wb['Schedule']
    

    # -------------------------
    # Colors
    # -------------------------
    section_colors = {'er-1': 'FFF2CC', 'er-2': 'F4CCCC', 'ew': 'D9D2E9'}
    neon_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # weekends
    light_blue = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')  # WR
    header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='Times New Roman', bold=True, size=12)
    regular_font = Font(name='Times New Roman', size=11)

    # -------------------------
    # Style header row
    # -------------------------
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # -------------------------
    # Style data rows
    # -------------------------
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        day_value = row[1].value  # Day column (assume 2nd col)
        is_weekend = day_value in ['Fri', 'Sat']

        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            # Match section prefix
            section = None
            for sec in section_colors:
                if col_name.startswith(sec):
                    section = sec
                    break

            if is_weekend:
                if col_name in ['wr', 'wr-1', 'wr-2']:
                    cell.fill = light_blue
                else:
                    cell.fill = neon_yellow

            elif section:
                cell.fill = PatternFill(start_color=section_colors[section],
                                        end_color=section_colors[section],
                                        fill_type='solid')
            # Font, alignment, border
            cell.font = regular_font
            cell.alignment = center_align
            cell.border = thin_border

    # -------------------------
    # Adjust column widths
    # -------------------------
    for i, col_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        header = col_cells[0].value
        if header == "date":
            width = 10
        elif header == "day":
            width = 8
        else:
            width = min(max_length + 6, 25)
        ws.column_dimensions[get_column_letter(i)].width = width

    # -------------------------
    # Split WR column into two if it contains multiple names
    # -------------------------


    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ Schedule saved as '{output_path}'")


# -------------------------
# Save Score DataFrame
# -------------------------
def save_score_as_excel(df, output_path="Score.xlsx"):
    """Save score DataFrame to Excel with gradient coloring on Score and Total Shifts columns."""

    # Normalize column names (remove spaces/strip) for safe access
    original_columns = df.columns.tolist()
    df = df.rename(columns=lambda x: x.strip().replace(" ", "_"))

    # Save DataFrame to Excel
    df_out = df.copy()
    df_out.columns = original_columns
    df_out.to_excel(output_path, index=False, sheet_name='Score')

    wb = load_workbook(output_path)
    ws = wb['Score']

    # Define borders and header style
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    header_font = Font(name='Times New Roman', bold=True, size=12, color='000000')
    regular_font = Font(name='Times New Roman', size=11)

    # Style header row
    for col_num, column in enumerate(original_columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Gradient helper function
    # Produces a red→yellow→green gradient based on ratio
    def get_rgb_from_ratio(ratio: float):
        ratio = max(0, min(ratio, 1))  # clamp between 0 and 1
        if ratio <= 0.5:
            # From red (low) to yellow (mid)
            r, g = 255, int(255 * (ratio / 0.5))
        else:
            # From yellow (mid) to green (high)
            g, r = 255, int(255 * (1 - (ratio - 0.5) / 0.5))
        return f"FF{r:02X}{g:02X}00"

    # Apply gradient to Score column
    if 'score' in df.columns and 'max_points' in df.columns:
        score_col_idx = original_columns.index('score') + 1
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            score_val = row.Score
            max_val = row.Max_Points
            ratio = 0 if max_val == 0 else score_val / max_val
            hex_color = get_rgb_from_ratio(ratio)
            cell = ws.cell(row=row_idx, column=score_col_idx)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell.font = regular_font
            cell.border = thin_border

    # Apply gradient to Total Shifts column
    if 'total_Shifts' in df.columns and 'max_Shifts' in df.columns:
        shifts_col_idx = original_columns.index('total shifts') + 1
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            shifts_val = row.Total_Shifts
            max_val = row.Max_Shifts
            ratio = 0 if max_val == 0 else shifts_val / max_val
            hex_color = get_rgb_from_ratio(ratio)
            cell = ws.cell(row=row_idx, column=shifts_col_idx)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
            cell.font = regular_font
            cell.border = thin_border

    # Adjust column widths dynamically
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 6, 25)

    # Freeze header row
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ Score saved as '{output_path}' with gradients on Score and Total Shifts")